#!/usr/bin/env python3

"""
Converts a specific sheet from an XLSX file to a CSV file,
preserving bold formatting as HTML <b> tags and saving cell values (not formulas).
"""

import argparse
import csv
from pathlib import Path
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles.fonts import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock 
from typing import Any, Union
from rich.console import Console
from tools.printer import printer as pr

console = Console()

def parse_arguments() -> argparse.Namespace:
    """Parses command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Convert an XLSX sheet to CSV, preserving bold as HTML <b> tags."
    )
    parser.add_argument(
        "input_doc_path",
        type=Path,
        help="Path to the input XLSX file."
    )
    parser.add_argument(
        "output_csv_path",
        type=Path,
        help="Path to the output CSV file."
    )
    parser.add_argument(
        "sheet_name",
        type=str,
        help="Name of the sheet to convert."
    )
    return parser.parse_args()

def format_cell_value(cell: Cell) -> str:
    """
    Formats the cell value, converting bold text to HTML <b> tags.
    Handles both entirely bold cells and rich text with mixed formatting.
    Converts whole number floats (e.g., 1.0) to integer strings (e.g., "1").
    """
    value: Any = cell.value # This will be the calculated value due to data_only=True
    font: Union[Font, None] = cell.font

    if value is None:
        return ""

    # Use CellRichText based on user feedback
    if isinstance(value, CellRichText): 
        formatted_parts: list[str] = []
        for part in value: 
            if isinstance(part, TextBlock):
                text_content: str = part.text if part.text is not None else ""
                if hasattr(part, 'font') and part.font and part.font.b:
                    formatted_parts.append(f"<b>{text_content}</b>")
                else:
                    formatted_parts.append(text_content)
            elif isinstance(part, str): # Plain string segments within CellRichText
                formatted_parts.append(part)
        return "".join(formatted_parts)
    
    # For non-rich text cells, or if data_only=True simplified it
    processed_value: str
    if isinstance(value, float) and value.is_integer():
        # If the value is a float but represents a whole number (e.g., 1.0, 2.0)
        # convert it to an integer string (e.g., "1", "2")
        processed_value = str(int(value))
    else:
        # Otherwise (it's a string, an int, or a float with actual decimals like 1.5)
        # convert it to string as is.
        processed_value = str(value)
    # ---- END OF MODIFICATION FOR NUMBER HANDLING ----
    
    # Now, apply bold formatting if needed, using the processed_value
    if font and font.b: # Check for bold on the original cell's font
        return f"<b>{processed_value}</b>"
    
    # If not bold, return the processed_value directly
    return processed_value

def convert_xlsx_to_csv(input_path: Path, output_path: Path, sheet_name: str) -> None:
    """
    Reads an XLSX sheet, formats cell content for bold text, and writes to a CSV file.
    Loads cell values, not formulas.
    """
    try:
        # Load workbook with rich_text=True for formatting and data_only=True for values
        workbook = openpyxl.load_workbook(input_path, rich_text=True, data_only=True)
    except FileNotFoundError:
        console.print(f"[red]Error: Input file not found at '{input_path}'")
        return
    except Exception as e:
        console.print(f"[red]Error loading workbook '{input_path}': {e}")
        return

    if sheet_name not in workbook.sheetnames:
        console.print(f"[red]Error: Sheet '{sheet_name}' not found in the workbook.")
        console.print(f"[green]Available sheets: {', '.join(workbook.sheetnames)}")
        return
    
    sheet = workbook[sheet_name]
    
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
            # When data_only=True, sheet.iter_rows() will yield cells with calculated values
            for row_idx, row_cells in enumerate(sheet.iter_rows()):
                try:
                    csv_writer.writerow([format_cell_value(cell) for cell in row_cells])
                except Exception as e_row:
                    console.print(f"[red]Error processing row {row_idx + 1} in sheet '{sheet_name}': {e_row}")
        console.print(f"[green]Successfully converted sheet '{sheet_name}' from '{input_path}' to '{output_path}' (values only).")

    except Exception as e:
        console.print(f"[red]Error writing CSV file to '{output_path}': {e}")

if __name__ == "__main__":
    pr.tic()
    console.print("[yellow]Converting XLSX to CSV...")
    console.print("[yellow]This script preserves bold formatting as HTML <b> tags.")
    args = parse_arguments()
    convert_xlsx_to_csv(args.input_doc_path, args.output_csv_path, args.sheet_name)
    pr.toc()

